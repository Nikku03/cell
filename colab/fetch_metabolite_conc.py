"""FETCH measured absolute metabolite concentrations (Problem 5, MEASURED tier).

Upgrades the metabolite layer from mostly-estimated toward measured. Primary source:
  Park, Chen, Rabinowitz et al. 2016 (Nat Chem Biol) — absolute intracellular concentrations (uM) measured
  by isotope-ratio LC-MS in mammalian iBMK cells, ~121 metabolites, each row carrying the full metabolite
  NAME and a BiGG abbreviation (nad/atp/pyr/g6p/...) -> near drop-in for Human-GEM.

Honesty: these are MEASURED facts (mammalian iBMK, not the specific human cell), so they anchor the model as
tier 'measured-literature(Park2016)'. We extract the numeric VALUES (facts) at runtime; we do not redistribute
the source file (it is (c) Springer Nature). Set MET_CONC_URL to override, or MET_CONC_URL="" to skip.

-> metabolite_conc_measured.tsv (metabolite  conc_uM  bigg  source)   [consumed by compute_metabolite_conc.py]
"""
import os, io, urllib.request
from pathlib import Path
OUT=Path("outputs/orphan"); H=Path("data/external_data/human")
PARK_URL=os.environ.get("MET_CONC_URL",
    "https://static-content.springer.com/esm/art%3A10.1038%2Fnchembio.2077/MediaObjects/"
    "41589_2016_BFnchembio2077_MOESM585_ESM.xlsx")

def _rows_from_xlsx(data):
    """yield (metabolite_name, concentration_M, bigg) from the Park MOESM585 sheet. Tries openpyxl then pandas."""
    try:
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws=wb[wb.sheetnames[0]]
        for r in ws.iter_rows(min_row=3, values_only=True):
            if r and len(r)>=5: yield r[1], r[2], r[4]
        return
    except Exception:
        pass
    import pandas as pd                                   # fallback
    df=pd.read_excel(io.BytesIO(data), header=1)
    for _,r in df.iterrows():
        yield r.get("Metabolite"), r.get("Concentration (M)"), r.get("Abbreviation")

def main():
    OUT.mkdir(parents=True, exist_ok=True)
    if not PARK_URL:
        print("  MET_CONC_URL empty -> measured metabolite concentrations skipped"); return
    cache=H/"park2016_conc.xlsx"
    try:
        if cache.exists() and cache.stat().st_size>10000:
            data=cache.read_bytes()
        else:
            print("  fetching Park 2016 absolute metabolite concentrations ...")
            data=urllib.request.urlopen(PARK_URL, timeout=60).read()
            try: H.mkdir(parents=True, exist_ok=True); cache.write_bytes(data)     # cache the source (not committed)
            except Exception: pass
    except Exception as e:
        print("  metabolite-concentration fetch failed:",repr(e)[:100],"-> layer stays thermodynamic-only"); return
    seen={}                                               # metabolite name -> (uM, bigg); constant per metabolite
    for name, conc_M, bigg in _rows_from_xlsx(data):
        if not name or conc_M in (None,""): continue
        try: uM=float(conc_M)*1e6
        except (TypeError,ValueError): continue
        if uM>0 and name.strip() not in seen:
            seen[name.strip()]=(round(uM,4), (str(bigg).strip() if bigg else ""))
    if not seen:
        print("  Park 2016: no concentrations parsed -> skipped"); return
    with open(OUT/"metabolite_conc_measured.tsv","w") as o:
        o.write("metabolite\tconc_uM\tbigg\tsource\n")
        for name,(uM,bigg) in sorted(seen.items()):
            o.write(f"{name}\t{uM}\t{bigg}\tPark2016\n")
    print(f"measured metabolite concentrations: {len(seen)} absolute values (uM) -> metabolite_conc_measured.tsv "
          f"(Park 2016, mammalian iBMK; anchors Problem 5 at measured-literature tier)")

if __name__=="__main__":
    main()
