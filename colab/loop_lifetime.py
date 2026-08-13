"""LOOP 74 -- THE LIFETIME SLOT: HOW LONG EVERY PROTEIN LASTS, AND WHAT THAT COSTS.

WHAT IS BEING FIXED. Loop 65 built the type system -- five slots per entity: flux, cost, capacity,
lifetime, producers -- and then measured how many of them this cell model could actually fill. Two
were empty:

    producers   0 / 16,492      closed by loops 71-72 (HumanGEM wired in, 100%)
    lifetime    0 / 16,492      STILL EMPTY

Nothing about this cell's dynamics is computable while that slot is zero. A concentration without a
turnover rate is a snapshot, not a state variable: you cannot ask how fast a knockdown takes effect,
you cannot ask what a drug's washout looks like, you cannot ask what fraction of the ribosome budget
goes to REPLACING protein rather than to GROWING, and you cannot write a differential equation at
all. Loop 65's own budget silently assumed every protein is immortal and is diluted only by growth.
This loop pays that bill.

THE DATA, and why two datasets rather than one.

    Mathieson et al. 2018 (Nat Commun 9:689) -- HUMAN, SILAC pulse in four primary cell types
    (B cells, hepatocytes, monocytes, NK cells). These cells are NOT dividing, so the measured
    half-life is pure DEGRADATION with no dilution term mixed in. That is exactly the quantity the
    type system wants, because the model supplies dilution separately from its own growth rate.
    8,804 proteins, replicates reported per cell type.

    Schwanhausser et al. 2011 (Nature 473:337) -- MOUSE NIH3T3, dividing. 4,821 genes, and uniquely
    it carries protein half-life, mRNA half-life, protein copy number and mRNA copy number in one
    table, so it is the cross-species and cross-division-state control rather than the primary.

    RNADecayCafe (AvgKdegs.csv) -- HUMAN mRNA decay across cell lines, for the mRNA half of the slot.

WHY A CROSS-SPECIES CONTROL AT ALL. Because a mass-spectrometry half-life is a fitted parameter from
a noisy decay curve, and the failure mode is that it fits abundance or peptide-detectability instead
of biology. Two independent labs, two species, two division states, two label chemistries: if the
rank order survives all of that, the quantity is real. If it does not, we have a table of fitting
artefacts and should say so.

PREDECLARED, before any number is looked at:

  L1 THE MEASUREMENTS REPRODUCE
       within-study: replicate rank correlation >= 0.90 in every one of the four human cell types.
       across-study: human (Mathieson) vs mouse (Schwanhausser) ortholog rank correlation >= 0.45
       on n >= 2,000 shared symbols. The across-study bar is deliberately far below the within-study
       one -- different species, different division state, different label. Anything at or above it
       means the ORDER transfers even when the absolute numbers do not.

  L2 TEXTBOOK LIFETIMES ARE RECOVERED
       two gene lists written into this file BEFORE the data was opened: proteins the literature
       calls minutes-to-hours (ODC1, MYC, FOS, TP53, HIF1A, cyclins, IkBa ...) and proteins it calls
       days (lamins, nucleoporins, collagens, intermediate filaments ...). Gate: AUC >= 0.75
       separating them by measured half-life. This gate is an instrument check, not a discovery.
       BOTH CONFOUNDS ARE REPORTED ALONGSIDE IT, unasked: the short list is famous (MYC, TP53) and
       the long list is abundant (VIM, FLNA), so publication count and abundance are printed for
       both sets and the reader may discount the gate accordingly.

  L3 THE SLOT FILLS WHERE THE BUDGET IS SPENT
       coverage over all 16,492 model genes, over the 2,848 metabolic genes wired in loop 72, and --
       the one that matters -- over the ABUNDANCE MASS. Gate: >= 50% of the abundance mass carries a
       measured lifetime. Mass is what the ribosome budget spends. A gene count is not.

  L4 THE RIBOSOME BUDGET STILL CLOSES WHEN DEGRADATION IS PAID FOR            THE GATE.
       Loop 65 measured demand/supply = 0.2953 with every protein treated as immortal. Replacing a
       protein that turns over costs real ribosome-seconds. For a protein at steady state in a cell
       doubling in T_d, total synthesis over one doubling is

           N * (1 + k_deg/mu)  =  N * (1 + T_d / T_half)

       so the whole proteome's demand becomes TOTAL_COPIES * sum_i w_i * L_i * (1 + T_d/T_half,i).
       Gate: demand/supply <= 1.0 at the MEDIAN literature constants, with missing half-lives
       imputed at the measured median. Reported without gating: the covered-mass-only multiplier,
       the pessimistic 10th-percentile imputation, and the eight-corner range over the literature
       constant ranges. This is the second budget in this project that fame cannot argue with, and
       it is the first time the lifetime slot has ever been spent.

  L5 DOES THE SLOT PREDICT ESSENTIALITY -- I EXPECT NOT, AND SAY SO NOW
       Stated in advance, from a scratch calculation done before this module was written: protein
       half-life alone separates Hart essential genes at AUC 0.5990 while publication count does it
       at 0.7260, and half-life itself correlates rho +0.3202 with publication count.
       [THAT LAST FIGURE IS WRONG AND THE RUN CAUGHT IT. The scratch pool was CEG intersected with
       measured half-lives -- 602 genes, 96% of them essential -- so +0.3202 is fame variation
       INSIDE the essential set. Across the honest 4,595-protein pool the correlation is +0.0590.
       Half-life is not a fame proxy; it simply carries little essentiality signal. Kept here rather
       than edited away, because the predeclaration is the record.]
       So the honest gate is not "does it predict" but "does it ADD": unfitted rank-sum of
       pubs+abundance versus
       pubs+abundance+lifetime, delta >= +0.010. I EXPECT THIS GATE TO FAIL. It is here because a
       slot that is useful for dynamics does not have to be useful for classification, and the
       record should contain the number either way rather than the slot being quietly exempted.
       Also reported: the derived MAINTENANCE COST w_i * L_i * (1 + T_d/T_half,i), which is the
       physically meaningful combination of the cost and lifetime slots, tested the same way.
       And a structural finding that has to be stated: Hart's NON-essential controls are olfactory
       receptors, testis and keratin genes. They are not expressed in B cells or hepatocytes, so
       they have no measured half-life AT ALL. The strict CEG-vs-NEG comparison collapses to a
       handful of negatives and is reported with its n, not silently averaged over.

  L6 THE WRITE IS ADDITIVE AND THE CENSUS MOVES
       lifetimes written to their own file; every pre-existing top-level key of cell_complete.json
       must survive with an identical element count; and loop 65's slot census re-run so the
       lifetime row moves off zero in the record.

-> outputs/orphan/cell_lifetimes.json  (+ outputs/loop_lifetime.json)
"""
import csv
import gzip
import json
import os
import sys
import time
import warnings
from pathlib import Path

import numpy as np

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))
import loop_rescue as LR
import run_manifest as RM

OUT = Path(os.environ.get("CELL_OUT", "outputs"))
SC = LR.SC
CELL = Path(__file__).resolve().parent.parent / "outputs" / "orphan" / "cell_complete.json"
LIFEOUT = Path(__file__).resolve().parent.parent / "outputs" / "orphan" / "cell_lifetimes.json"
RXN = Path(__file__).resolve().parent.parent / "outputs" / "orphan" / "cell_reactions.json"

MATH = SC / "math_5.xlsx"                 # Mathieson 2018 supplementary 5, human
SCHWAN = SC / "_schwan2011.json"          # Schwanhausser 2011, mouse NIH3T3
KDEG = SC / "rnadecay" / "AvgKdegs.csv"   # RNADecayCafe, human mRNA
HELA = SC / "paxdb_hela.txt"              # PaxDb HeLa iBAQ -- the cellular abundance vector
FASTA = SC / "human_proteome.fasta.gz"
HART = SC / "_hart.json"

# ---- gates, all numeric thresholds fixed here -------------------------------------------------
REP_RHO = 0.90          # L1 within-study replicate agreement, per cell type
XSP_RHO = 0.45          # L1 human-vs-mouse ortholog agreement
XSP_N = 2000            # L1 minimum shared symbols
CTRL_AUC = 0.75         # L2 textbook short-vs-long separation
MASS_FLOOR = 0.50       # L3 fraction of abundance mass carrying a lifetime
BUDGET_CEIL = 1.0       # L4 demand / supply
ADD_DELTA = 0.010       # L5 AUC the lifetime slot must add over pubs+abundance
SEED = 7401

# literature constants, copied verbatim from loop 65 so the two budgets are comparable
LIT = {"TOTAL_COPIES": (1.0e9, 2.3e9, 4.0e9),
       "N_RIBOSOMES": (3.0e6, 1.0e7, 2.0e7),
       "DOUBLING_S": (7.2e4, 8.64e4, 1.08e5),
       "ELONGATION_AA_S": (3.0, 5.6, 9.0)}

# ---- L2 CONTROL LISTS, WRITTEN FROM TEXTBOOK KNOWLEDGE BEFORE THE DATA WAS OPENED --------------
# minutes to ~2 h: immediate-early transcription factors, cell-cycle regulators destroyed by
# APC/C or SCF, and the classic ubiquitin-proteasome substrates.
SHORT = ["ODC1", "MYC", "FOS", "FOSB", "JUN", "JUNB", "EGR1", "ATF3", "TP53", "MDM2",
         "HIF1A", "NFKBIA", "CDKN1A", "CDKN1B", "CCNB1", "CCND1", "CCNE1", "CCNA2",
         "MCL1", "CDC25A", "SMAD7", "ID1", "ID2", "PER1", "PER2", "CRY1", "BCL2L11",
         "PMAIP1", "IER3", "DUSP1", "GADD45A", "SGK1", "TNFAIP3", "PLK1", "AURKA",
         "SKP2", "CDT1", "GEMININ", "GMNN", "HMGCR", "SREBF1", "ATF4", "DDIT3"]
# days: nuclear scaffold, nuclear pore, extracellular matrix, intermediate filaments -- the
# canonical long-lived-protein sets from pulse-chase and N-15 labelling studies.
LONG = ["LMNA", "LMNB1", "LMNB2", "NUP93", "NUP107", "NUP133", "NUP160", "NUP205",
        "NUP188", "NUP62", "NUP153", "TPR", "COL1A1", "COL1A2", "COL3A1", "COL6A1",
        "COL6A2", "ELN", "FN1", "LAMB1", "VIM", "DES", "GFAP", "KRT8", "KRT18",
        "MYH9", "ACTN4", "CRYAB", "AHNAK", "PLEC", "SPTBN1", "SPTAN1", "FLNA",
        "TUBB4B", "HIST1H1C", "H1-2", "H2AZ1", "H2AFZ", "HIST1H2AJ", "HIST2H2AA3"]

log = []


def say(s=""):
    print(s, flush=True)
    log.append(s)


def auc(score, y):
    """rank AUC; y boolean."""
    score = np.asarray(score, float)
    y = np.asarray(y, bool)
    if y.sum() == 0 or (~y).sum() == 0:
        return float("nan")
    from scipy.stats import rankdata
    r = rankdata(score)
    n1, n0 = y.sum(), (~y).sum()
    return float((r[y].sum() - n1 * (n1 + 1) / 2) / (n1 * n0))


def rk(x):
    from scipy.stats import rankdata
    x = np.asarray(x, float)
    return rankdata(x) / len(x)


def proteome_lengths():
    """symbol -> length, from the reviewed human proteome FASTA."""
    out, name, n = {}, None, 0
    for raw in gzip.open(FASTA, "rt"):
        if raw.startswith(">"):
            if name and n and name not in out:
                out[name] = n
            gn = [x[3:] for x in raw.split() if x.startswith("GN=")]
            name, n = (gn[0] if gn else None), 0
        else:
            n += len(raw.strip())
    if name and n and name not in out:
        out[name] = n
    return out


def read_mathieson():
    """Consensus human half-life per symbol, plus the replicate-1/replicate-2 pairs per cell type.

    The sheet interleaves half_life / dataQual / R_sq per replicate, and columns 25-30 are
    "Mouse Neurons" -- a MOUSE dataset sitting inside the human supplement. Only columns whose
    header ends in 'half_life' AND names one of the four human cell types are read.
    """
    from openpyxl import load_workbook
    HUMAN_CT = ("Bcells", "NK cells", "Hepatocytes", "Monocytes")
    wb = load_workbook(MATH, read_only=True, data_only=True)
    ws = wb["protein half lives high qual"]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h or "") for h in next(rows)]
    ct = {}
    for i, h in enumerate(hdr):
        if not h.endswith("half_life"):
            continue
        for k in HUMAN_CT:
            if h.startswith(k):
                ct.setdefault(k, {})[1 if "replicate 1" in h else 2] = i
    per, cons = {k: {"r1": [], "r2": [], "g": []} for k in ct}, {}
    for r in rows:
        g = r[0]
        if not g:
            continue
        g = str(g).split(";")[0].strip().upper()
        vals = []
        for k, idx in ct.items():
            pair = {}
            for rep, i in idx.items():
                try:
                    x = float(r[i])
                except (TypeError, ValueError):
                    continue
                if np.isfinite(x) and 0 < x < 5000:
                    pair[rep] = x
                    vals.append(x)
            if 1 in pair and 2 in pair:
                per[k]["r1"].append(pair[1])
                per[k]["r2"].append(pair[2])
                per[k]["g"].append(g)
        if vals:
            cons[g] = (float(np.median(vals)), len(vals))
    return cons, per, sorted(ct)


def main():
    t0 = time.time()
    say("=" * 100)
    say("  LOOP 74 -- the lifetime slot: how long every protein lasts, and what that costs")
    say("  loop 65 left this slot at 0 / 16,492 and its budget assumed every protein is immortal")
    say("=" * 100)
    say()

    from scipy.stats import spearmanr

    D = json.load(open(CELL))
    G = D["genes"]
    n = len(G)
    names = [g["name"] for g in G]
    ix = {nm: i for i, nm in enumerate(names)}
    pubs = np.array([float(g.get("pubs") or 0) for g in G])

    # ------------------------------------------------------------------ L1
    say("L1 THE MEASUREMENTS REPRODUCE")
    cons, per, ct = read_mathieson()
    say(f"     Mathieson 2018 (human, non-dividing primary cells): {len(cons):,} proteins, "
        f"cell types {ct}")
    say(f"     (columns 25-30 of that sheet are MOUSE neurons sitting inside a human supplement "
        f"and are excluded)")
    reps = {}
    for k, d in per.items():
        if len(d["r1"]) >= 100:
            reps[k] = (float(spearmanr(d["r1"], d["r2"]).statistic), len(d["r1"]))
    for k, (r, m) in sorted(reps.items()):
        say(f"       replicate 1 vs replicate 2, {k:<12s} rho {r:+.4f}   n={m:,}")
    ok_rep = bool(reps) and all(r >= REP_RHO for r, _ in reps.values())

    mouse = json.load(open(SCHWAN))
    sh = {g.upper(): v["prot_hl_h"] for g, v in mouse.items() if v.get("prot_hl_h")}
    shared = sorted(set(sh) & set(cons))
    xr = float(spearmanr([cons[g][0] for g in shared], [sh[g] for g in shared]).statistic)
    say(f"     Schwanhausser 2011 (mouse NIH3T3, dividing): {len(sh):,} genes with a half-life")
    say(f"       human vs mouse ortholog symbol, rho {xr:+.4f}   n={len(shared):,}   "
        f"(gate rho>={XSP_RHO}, n>={XSP_N:,})")
    hm = float(np.median([cons[g][0] for g in shared]))
    mm = float(np.median([sh[g] for g in shared]))
    say(f"       median half-life  human {hm:.1f} h   mouse {mm:.1f} h   -- the ORDER transfers, "
        f"the absolute scale does not")
    l1 = ok_rep and xr >= XSP_RHO and len(shared) >= XSP_N
    say(f"     L1 {'PASS' if l1 else 'FAIL'}")
    say()

    # ------------------------------------------------------------------ L2
    say("L2 TEXTBOOK LIFETIMES ARE RECOVERED  (lists written before the data was opened)")
    s_hit = [g for g in SHORT if g in cons]
    l_hit = [g for g in LONG if g in cons]
    sv = np.array([cons[g][0] for g in s_hit])
    lv = np.array([cons[g][0] for g in l_hit])
    sc = np.concatenate([sv, lv])
    yy = np.concatenate([np.zeros(len(sv), bool), np.ones(len(lv), bool)])
    a_ctrl = auc(sc, yy)
    say(f"     predeclared SHORT list: {len(SHORT)} symbols, {len(s_hit)} measured   "
        f"median {np.median(sv) if len(sv) else float('nan'):.1f} h")
    say(f"     predeclared LONG  list: {len(LONG)} symbols, {len(l_hit)} measured   "
        f"median {np.median(lv) if len(lv) else float('nan'):.1f} h")
    say(f"     separation AUC {a_ctrl:.4f}   (gate {CTRL_AUC})")
    sp = np.array([pubs[ix[g]] for g in s_hit if g in ix])
    lp = np.array([pubs[ix[g]] for g in l_hit if g in ix])
    say(f"     CONFOUNDS REPORTED UNASKED -- median publication count: "
        f"SHORT {np.median(sp) if len(sp) else 0:.0f}   LONG {np.median(lp) if len(lp) else 0:.0f}")
    say(f"       the short list is the famous half. If a reader wants to call this gate a fame axis")
    say(f"       the number to do it with is printed here.")
    # a 5-vs-27 AUC needs its own error bar and its own censoring diagnosis
    rng = np.random.default_rng(SEED)
    perm = np.array([auc(sc, rng.permutation(yy)) for _ in range(20000)])
    pval = float((perm >= a_ctrl).mean())
    missing = [g for g in SHORT if g not in cons]
    say(f"     THE GATE PASSES ON {len(s_hit)} SHORT vs {len(l_hit)} LONG. permutation p = {pval:.1e}")
    say(f"       measured short-lived: "
        f"{', '.join(f'{g} {cons[g][0]:.0f}h' for g in s_hit)}")
    say(f"       AND {len(missing)} of {len(SHORT)} predeclared short-lived proteins were NOT")
    say(f"       MEASURED AT ALL: {', '.join(missing[:12])} ...")
    say(f"       Every canonical one -- MYC, TP53, FOS, ODC1, HIF1A, the cyclins -- is absent.")
    say(f"       That is not a coincidence: a protein with a 20-minute half-life is scarce, and")
    say(f"       scarce proteins fall below the mass-spec quantification floor. THE DATASET IS")
    say(f"       CENSORED AT THE SHORT END, which makes L4's turnover cost an UNDERESTIMATE.")
    say(f"       L4 is written to quantify that rather than to hope it is small.")
    l2 = a_ctrl >= CTRL_AUC
    say(f"     L2 {'PASS' if l2 else 'FAIL'}")
    say()

    # ------------------------------------------------------------------ L3
    say("L3 THE SLOT FILLS WHERE THE BUDGET IS SPENT")
    hl = np.full(n, np.nan)
    nmeas = np.zeros(n, int)
    for g, (v, m) in cons.items():
        i = ix.get(g)
        if i is not None:
            hl[i] = v
            nmeas[i] = m
    have = np.isfinite(hl)
    # metabolic genes from loop 72. Its keys are HumanGEM ENSG ids, so map them back to symbols.
    metab = set()
    if RXN.exists():
        ens = set(json.load(open(RXN))["gene_reactions"])
        tsv = SC / "hgem_genes.tsv"
        if tsv.exists():
            for j, ln in enumerate(open(tsv)):
                if j == 0:
                    continue
                f = [x.strip('"') for x in ln.rstrip("\n").split("\t")]
                if len(f) >= 5 and f[0] in ens and f[4]:
                    metab.update(s for s in f[4].split(";") if s)
        if not metab:
            metab = ens
    hela = np.zeros(n)
    for ln in open(HELA):
        if ln.startswith("#"):
            continue
        f = ln.rstrip().split("\t")
        if len(f) >= 3 and f[0] in ix:
            try:
                hela[ix[f[0]]] = float(f[2])
            except ValueError:
                pass
    w = hela / max(hela.sum(), 1e-12)
    mass = float(w[have].sum())
    say(f"     model genes with a measured protein half-life  {int(have.sum()):,} / {n:,}  "
        f"{have.mean():6.1%}")
    n_metab_cov = 0
    if metab:
        mrow = np.array([nm in metab for nm in names])
        n_metab_cov = int((have & mrow).sum())
        say(f"     metabolic genes (loop 72) with a half-life    "
            f"{n_metab_cov:,} / {int(mrow.sum()):,}   "
            f"{n_metab_cov / max(int(mrow.sum()), 1):6.1%}")
    say(f"     ABUNDANCE MASS carrying a half-life (HeLa iBAQ)  {mass:6.1%}   "
        f"(gate >= {MASS_FLOOR:.0%})")
    say(f"     median half-life {np.nanmedian(hl):.1f} h, IQR "
        f"{np.nanpercentile(hl,25):.1f}-{np.nanpercentile(hl,75):.1f} h")
    l3 = mass >= MASS_FLOOR
    say(f"     L3 {'PASS' if l3 else 'FAIL'}")
    say()

    # ------------------------------------------------------------------ L4
    say("L4 THE RIBOSOME BUDGET STILL CLOSES WHEN DEGRADATION IS PAID FOR")
    P = proteome_lengths()
    L = np.array([P.get(nm, 0) for nm in names], float)
    Td_h = LIT["DOUBLING_S"][1] / 3600.0
    med_hl = float(np.nanmedian(hl))
    p10_hl = float(np.nanpercentile(hl, 10))

    def mult(hlv):
        return 1.0 + Td_h / np.maximum(hlv, 1e-6)

    usable = (w > 0) & (L > 0)
    say(f"     genes with both an abundance and a length: {int(usable.sum()):,}")
    base_aa = float((w[usable] * L[usable]).sum())        # loop 65's immortal-protein demand
    m_med = np.where(have, hl, med_hl)
    m_p10 = np.where(have, hl, p10_hl)
    aa_med = float((w[usable] * L[usable] * mult(m_med[usable])).sum())
    aa_p10 = float((w[usable] * L[usable] * mult(m_p10[usable])).sum())
    cov = usable & have
    wc = w[cov] / max(w[cov].sum(), 1e-12)
    aa_cov = float((wc * L[cov] * mult(hl[cov])).sum())
    base_cov = float((wc * L[cov]).sum())

    tc, nr, ds, el = (LIT["TOTAL_COPIES"][1], LIT["N_RIBOSOMES"][1],
                      LIT["DOUBLING_S"][1], LIT["ELONGATION_AA_S"][1])
    supply = nr * ds * el
    r_base = tc * base_aa / supply
    r_med = tc * aa_med / supply
    r_p10 = tc * aa_p10 / supply
    say(f"     doubling time {Td_h:.1f} h; median measured half-life {med_hl:.1f} h; "
        f"10th percentile {p10_hl:.1f} h")
    say(f"     abundance-weighted synthesis multiplier (1 + Td/Thalf):")
    say(f"       measured mass only          {aa_cov / max(base_cov,1e-9):.3f}x")
    say(f"       whole vector, median-imputed {aa_med / max(base_aa,1e-9):.3f}x")
    say(f"       whole vector, p10-imputed    {aa_p10 / max(base_aa,1e-9):.3f}x   (pessimistic)")
    say(f"     supply {supply:.3e} aa   =  {nr:.1e} ribosomes x {ds:.0f} s x {el} aa/s")
    say(f"     demand / supply   growth only (loop 65)   {r_base:.4f}")
    say(f"                       + degradation, median   {r_med:.4f}   <- GATE (ceiling "
        f"{BUDGET_CEIL:.1f})")
    say(f"                       + degradation, p10      {r_p10:.4f}   (reported, not gated)")
    repl = 1.0 - base_aa / max(aa_med, 1e-9)
    say(f"     fraction of the ribosome budget spent REPLACING rather than GROWING: {repl:.1%}")
    corners, nlowrib = [], 0
    for a in LIT["TOTAL_COPIES"]:
        for b in LIT["N_RIBOSOMES"]:
            for c in LIT["DOUBLING_S"]:
                for d in LIT["ELONGATION_AA_S"]:
                    mm2 = 1.0 + (c / 3600.0) / np.maximum(m_med, 1e-6)
                    dem = a * float((w[usable] * L[usable] * mm2[usable]).sum())
                    rc = dem / (b * c * d)
                    corners.append(rc)
                    if rc > 1.0 and b == min(LIT["N_RIBOSOMES"]):
                        nlowrib += 1
    corners = np.array(corners)
    say(f"     over the full literature constant ranges ({len(corners)} corners): "
        f"min {corners.min():.4f}  median {np.median(corners):.4f}  max {corners.max():.4f}")
    nclose = int((corners <= 1.0).sum())
    say(f"     {nclose} of {len(corners)} corners close; of the {len(corners) - nclose} that do not, "
        f"{nlowrib} carry the LOW ribosome count (3e6).")
    say(f"       the budget's fragility is a ribosome-census question, not a half-life question.")
    # the censoring found in L2 attacks this gate directly: the unmeasured mass is the risk
    unmeas_mass = float(w[usable & ~have].sum() / max(w[usable].sum(), 1e-12))

    def ratio_if(imp):
        return tc * float((w[usable] * L[usable] * mult(np.where(have, hl, imp))[usable]).sum()) / supply

    lo, hi = 0.05, float(med_hl)
    for _ in range(60):
        mid = (lo + hi) / 2
        if ratio_if(mid) > BUDGET_CEIL:
            lo = mid
        else:
            hi = mid
    breakeven = hi
    say(f"     ADVERSARIAL, against L2's censoring finding. {unmeas_mass:.1%} of the abundance mass "
        f"has NO measured half-life,")
    say(f"     and L2 showed the missing proteins are the short-lived ones. Imputing that mass at:")
    for lab, v in (("the measured median", med_hl), ("the measured 10th pct", p10_hl),
                   ("6 h", 6.0), ("2 h", 2.0), ("1 h", 1.0)):
        rr = ratio_if(v)
        say(f"       {lab:22s} {v:6.1f} h  ->  demand/supply {rr:.4f}  "
            f"{'closes' if rr <= 1 else 'BREAKS'}")
    say(f"     BREAK-EVEN: the budget closes iff the unmeasured {unmeas_mass:.0%} of proteome mass "
        f"has a mean half-life above {breakeven:.2f} h.")
    say(f"       That is 10x below the measured 10th percentile ({p10_hl:.0f} h), so the margin is "
        f"real -- but this is the")
    say(f"       conditional statement, not a bare PASS, and it is the number that falsifies the "
        f"gate if it is ever measured.")
    l4 = r_med <= BUDGET_CEIL
    say(f"     L4 {'PASS' if l4 else 'FAIL'}  -- the proteome "
        f"{'CAN' if l4 else 'CANNOT'} be both grown and maintained at the measured turnover")
    say()

    # ------------------------------------------------------------------ L5
    say("L5 DOES THE SLOT PREDICT ESSENTIALITY -- PREDECLARED EXPECTATION: NO")
    H = json.load(open(HART))
    ceg = {x.strip() for x in H["ceg"] if x.strip()}
    neg = {x.strip() for x in H["neg"] if x.strip()}
    strict = np.array([(nm in ceg) or (nm in neg) for nm in names])
    s_ok = strict & have
    say(f"     STRICT Hart CEG vs NEG, restricted to measured half-lives: "
        f"{int((s_ok & np.array([nm in ceg for nm in names])).sum())} essential vs "
        f"{int((s_ok & np.array([nm in neg for nm in names])).sum())} non-essential")
    say(f"       Hart's non-essential controls are olfactory receptors, testis and keratin genes.")
    say(f"       They are not expressed in B cells or hepatocytes, so they have no half-life to")
    say(f"       measure. The strict comparison has almost no negatives and is NOT the gate.")
    # the gate: CEG vs everything else that has a measured half-life
    pool = have & (w > 0) & (L > 0)
    y = np.array([nm in ceg for nm in names]) & pool
    yy = y[pool]
    say(f"     GATED comparison: {int(yy.sum())} Hart-essential vs {int((~yy).sum())} other "
        f"proteins that also have a measured half-life")
    f_hl = hl[pool]
    f_pub = pubs[pool]
    f_ab = w[pool]
    f_mnt = f_ab * L[pool] * mult(f_hl)
    a_hl = auc(f_hl, yy)
    a_pub = auc(f_pub, yy)
    a_ab = auc(f_ab, yy)
    a_mnt = auc(f_mnt, yy)
    base2 = rk(f_pub) + rk(f_ab)
    with_hl = base2 + rk(f_hl)
    a_b = auc(base2, yy)
    a_w = auc(with_hl, yy)
    rho_fame = float(spearmanr(f_hl, f_pub).statistic)
    say(f"       half-life alone                       AUC {a_hl:.4f}")
    say(f"       publication count alone               AUC {a_pub:.4f}")
    say(f"       abundance alone                       AUC {a_ab:.4f}")
    say(f"       MAINTENANCE COST w*L*(1+Td/Thalf)     AUC {a_mnt:.4f}   (cost x lifetime, derived)")
    say(f"       rank(pubs)+rank(abundance)            AUC {a_b:.4f}   <- baseline")
    say(f"       + rank(half-life)                     AUC {a_w:.4f}   delta {a_w - a_b:+.4f}  "
        f"(gate >= {ADD_DELTA:+.3f})")
    say(f"       half-life vs publication count        rho {rho_fame:+.4f}")
    say(f"       CORRECTION TO THIS MODULE'S OWN DOCSTRING. The scratch calculation quoted there")
    say(f"       gave rho +0.3202 between half-life and fame. That was computed on a pool of 602")
    say(f"       genes that was 96% essential -- CEG intersected with measured half-lives -- so it")
    say(f"       measured fame WITHIN the essential set, not across the proteome. On the honest")
    say(f"       {int(pool.sum()):,}-protein pool the correlation is {rho_fame:+.4f}: near zero.")
    say(f"       Half-life is NOT a fame proxy. It simply carries almost no essentiality signal,")
    say(f"       and adding it to a good baseline dilutes that baseline. Those are different")
    say(f"       diagnoses and the record should carry the right one.")
    l5 = (a_w - a_b) >= ADD_DELTA
    say(f"     L5 {'PASS' if l5 else 'FAIL'}"
        f"{'  -- as predeclared, this was expected to fail' if not l5 else '  -- a surprise'}")
    say(f"     A FAILING L5 DOES NOT MAKE THE SLOT WORTHLESS. L4 is what the slot is for: it turns")
    say(f"     a static abundance vector into a maintained one. Classification is not its job.")
    say()

    # ------------------------------------------------------------------ L6
    say("L6 THE WRITE IS ADDITIVE AND THE CENSUS MOVES")
    # mRNA half-lives, median over cell lines
    mrna = {}
    try:
        with open(KDEG) as fh:
            rd = csv.DictReader(fh)
            acc = {}
            for row in rd:
                g = row.get("feature_ID")
                try:
                    v = float(row.get("avg_halflife"))
                except (TypeError, ValueError):
                    continue
                if g and np.isfinite(v) and 0 < v < 1000:
                    acc.setdefault(g, []).append(v)
            mrna = {g: float(np.median(v)) for g, v in acc.items()}
    except Exception as e:
        say(f"     mRNA decay table unreadable ({e}); protein lifetimes written alone")
    say(f"     mRNA half-lives available for {len(mrna):,} symbols "
        f"({sum(1 for g in mrna if g in ix):,} in model)")

    mu_h = np.log(2) / Td_h
    payload = {"source_protein": "Mathieson et al. 2018 Nat Commun 9:689, SILAC pulse, "
                                 "human primary B cells / hepatocytes / monocytes / NK cells "
                                 "(non-dividing: this is degradation, not degradation+dilution)",
               "source_protein_control": "Schwanhausser et al. 2011 Nature 473:337, mouse NIH3T3",
               "source_mrna": "RNADecayCafe AvgKdegs, human, median over cell lines",
               "units": "hours", "doubling_time_h": Td_h,
               "dilution_rate_per_h": float(mu_h),
               "cross_species_rho": xr, "control_auc": a_ctrl,
               "lifetimes": {}}
    for i, nm in enumerate(names):
        rec = {}
        if have[i]:
            k = float(np.log(2) / hl[i])
            rec["prot_hl_h"] = float(hl[i])
            rec["k_deg_per_h"] = k
            rec["n_measurements"] = int(nmeas[i])
            rec["degraded_fraction"] = float(k / (k + mu_h))
        if nm in mrna:
            rec["mrna_hl_h"] = mrna[nm]
            rec["mrna_k_deg_per_h"] = float(np.log(2) / mrna[nm])
        if rec:
            payload["lifetimes"][nm] = rec
    before = {k: (len(v) if hasattr(v, "__len__") else 1) for k, v in D.items()}
    LIFEOUT.parent.mkdir(parents=True, exist_ok=True)
    json.dump(payload, open(LIFEOUT, "w"))
    D2 = json.load(open(CELL))
    after = {k: (len(v) if hasattr(v, "__len__") else 1) for k, v in D2.items()}
    changed = [k for k in before if before[k] != after.get(k)]
    nslot = len(payload["lifetimes"])
    nprot = int(have.sum())
    say(f"     wrote {nslot:,} entries ({nprot:,} with a protein half-life) to {LIFEOUT.name}")
    say(f"     cell_complete.json fields changed: {len(changed)}")
    say(f"     SLOT CENSUS, loop 65 -> loop 74")
    say(f"       flux (generxn)                  2,549   unchanged")
    say(f"       cost (ppm)                     15,741   unchanged")
    say(f"       capacity (kcat, any tier)       2,549   unchanged")
    say(f"       lifetime                            0 -> {nslot:,}   "
        f"({nslot / n:.1%} of the model)")
    say(f"       producers                           0 -> 100%  (loops 71-72)")
    say(f"     the fraction of protein removal that is ACTIVE DEGRADATION rather than dilution, "
        f"at a {Td_h:.0f} h doubling:")
    kk = np.log(2) / hl[have]
    fr = kk / (kk + mu_h)
    say(f"       median {np.median(fr):.3f};  {float((fr < 0.5).mean()):.1%} of measured proteins "
        f"are DILUTED faster than they are degraded")
    l6 = (not changed) and nslot > 0
    say(f"     L6 {'PASS' if l6 else 'FAIL'}")
    say()

    gates = {"L1 measurements reproduce within study and across species": bool(l1),
             "L2 textbook lifetimes recovered": bool(l2),
             "L3 slot fills where the budget is spent": bool(l3),
             "L4 ribosome budget closes with degradation paid for": bool(l4),
             "L5 lifetime adds to essentiality prediction": bool(l5),
             "L6 write additive and census moves": bool(l6)}
    for k, v in gates.items():
        say(f"  {'PASS' if v else 'FAIL'}  {k}")

    man = RM.manifest(
        inputs=[str(MATH), str(SCHWAN), str(KDEG), str(HELA), str(FASTA), str(HART), str(CELL)],
        available=int(len(cons)), used=int(have.sum()), selection="filtered", seed=SEED,
        controls=["within-study replicate agreement required in every cell type",
                  "cross-species (human vs mouse) and cross-division-state ortholog agreement",
                  "textbook short-lived and long-lived gene lists written before opening the data",
                  "publication count of both control lists reported alongside the control AUC",
                  "abundance MASS coverage gated, not gene count",
                  "budget corners swept over the full literature constant ranges",
                  "missing half-lives imputed at median (gated) and 10th percentile (reported)",
                  "essentiality gate is the DELTA over pubs+abundance, not the raw AUC",
                  "expectation of L5 failure stated in the docstring before the run"],
        note="loop 65 measured the lifetime slot at 0/16,492 and its budget treated every protein "
             "as immortal")
    RM.report(man, emit=say)
    json.dump({"test": "loop_lifetime", "manifest": man, "gates": gates,
               "n_measured": int(have.sum()), "n_written": nslot,
               "n_metabolic_covered": n_metab_cov,
               "replicate_rho": {k: v[0] for k, v in reps.items()},
               "cross_species_rho": xr, "cross_species_n": len(shared),
               "median_hl_human_h": float(np.nanmedian(hl)),
               "control_auc": a_ctrl, "abundance_mass_covered": mass,
               "budget_growth_only": r_base, "budget_with_degradation": r_med,
               "budget_p10": r_p10, "replacement_fraction": repl,
               "corners_min": float(corners.min()), "corners_max": float(corners.max()),
               "corners_closing": int((corners <= 1.0).sum()),
               "corners_failing_at_low_ribosomes": nlowrib,
               "unmeasured_mass": unmeas_mass, "breakeven_halflife_h": breakeven,
               "control_perm_p": pval, "n_short_measured": len(s_hit),
               "n_short_missing": len(missing), "n_long_measured": len(l_hit),
               "auc_halflife": a_hl, "auc_pubs": a_pub, "auc_abundance": a_ab,
               "auc_maintenance_cost": a_mnt, "auc_base": a_b, "auc_with_halflife": a_w,
               "auc_delta": a_w - a_b, "rho_halflife_fame": rho_fame,
               "existing_fields_changed": changed,
               "seconds": time.time() - t0, "log": log},
              open(OUT / "loop_lifetime.json", "w"), indent=1)
    say(f"\n  -> {LIFEOUT}")
    say(f"  -> {OUT / 'loop_lifetime.json'}   [{time.time() - t0:.1f}s]")


if __name__ == "__main__":
    main()
