"""ComplexAssemblyDetector: predict essentiality from known-complex membership.

Rationale: the existing trajectory-based detectors (ShortWindow, PerRule,
EnsembleDetector, RedundancyAware) all bottom out around MCC ~0.12-0.15
on Breuer 2019 because ~15 of the 40 balanced-panel essentials are
non-catalytic — ribosomal proteins, tRNA synthetases, translation
factors, transcription machinery. These genes don't catalyse a metabolic
rule, so a rule-silencing detector cannot see them. Pool-deviation
detectors also miss them because central metabolites ride out short
simulations on stored pools.

But they ARE subunits of known multi-protein complexes whose assembly
fails without them. The ``complex_formation.xlsx`` data under
``cell_sim/data/Minimal_Cell_ComplexFormation/input_data/`` — the same
file the real simulator already consumes for its protein-complex
dynamics — encodes 25 complexes covering 121 genes, including all 58
ribosomal subunits, RNA polymerase, SecYEGDF, ATP synthase, primary
ABC transporters, etc. 87 of those 121 genes carry Breuer's
'Essential' label (72 %) and only 3 are 'Nonessential'.

A purely static "is the knocked-out gene a subunit of an active
complex?" lookup yields:

  * 87 true positives (essentials correctly called)
  *  3 false positives (Breuer nonessentials in complex list)
  * 25 quasiessentials flagged (mixed with the essentials in scoring)
  *  6 locus_tags not in the Breuer table (unknown)

This is a *structural* detector: it requires no simulator run and no
trajectory comparison. Each predict() call is a constant-time dict
lookup. Designed to be composed with the trajectory-based detectors:
when the structural signal fires it is high-precision, and the
trajectory detectors handle genes outside the known-complex set.

Offline-first: the only input is the bundled .xlsx file, parsed once
at construction and cached in memory. No network or API calls.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional

from cell_sim.layer6_essentiality.harness import FailureMode, Trajectory


_DEFAULT_COMPLEX_XLSX = (
    Path(__file__).resolve().parents[1]
    / "data" / "Minimal_Cell_ComplexFormation"
    / "input_data" / "complex_formation.xlsx"
)


def _normalise_locus(raw: str) -> str:
    """Take a raw identifier like "0685", "JCVISYN3A_0685" or 685 and
    return the canonical locus_tag "JCVISYN3A_0685"."""
    s = str(raw).strip()
    if s.startswith("JCVISYN3A_"):
        return s
    # Strip any leading non-digits (e.g. "G_MMSYN1_0685" if the sheet uses
    # the SBML-compiled id).
    digits = "".join(c for c in s if c.isdigit())
    if not digits:
        return s
    return f"JCVISYN3A_{digits.zfill(4)}"


@dataclass(frozen=True)
class ComplexMembership:
    complex_name: str
    stoichiometry: int
    init_count: int
    pdb_structures: tuple[str, ...] = ()

    @property
    def is_active(self) -> bool:
        """True when the simulator is configured with >0 copies of
        this complex — a knockout of any subunit should then prevent
        the cell from maintaining the pool."""
        return self.init_count > 0


@dataclass
class ComplexAssemblyKB:
    """Knowledge base of gene -> complex memberships parsed from
    complex_formation.xlsx. Cached on the class; loading the 25-row
    sheet is fast but we memoise so a sweep with hundreds of
    predict() calls doesn't re-hit openpyxl on every one."""

    gene_to_memberships: dict[str, list[ComplexMembership]] = field(default_factory=dict)

    @classmethod
    def load(cls, path: Optional[Path] = None) -> "ComplexAssemblyKB":
        path = Path(path) if path else _DEFAULT_COMPLEX_XLSX
        if not path.exists():
            raise FileNotFoundError(
                f"Complex-formation spreadsheet not found at {path}. "
                "This is a bundled offline data file; check the repo "
                "checkout is complete."
            )
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Complexes" not in wb.sheetnames:
            raise ValueError(f"{path} has no 'Complexes' sheet")
        ws = wb["Complexes"]

        kb: dict[str, list[ComplexMembership]] = {}
        for row in list(ws.iter_rows(values_only=True))[1:]:
            name, genes_raw, stoich_raw, init_raw, _path, pdb_raw, _how = (
                (row + (None,) * 7)[:7]
            )
            if not name or not genes_raw:
                continue
            genes = [g.strip() for g in str(genes_raw).split(";") if g.strip()]
            stoich = [int(s.strip()) for s in str(stoich_raw).split(";") if s.strip()]
            try:
                init_count = int(init_raw) if init_raw is not None else 0
            except (TypeError, ValueError):
                init_count = 0
            pdbs = tuple(
                p.strip()
                for p in (str(pdb_raw) if pdb_raw else "").replace(",", ";").split(";")
                if p.strip()
            )
            if len(stoich) < len(genes):
                stoich = stoich + [1] * (len(genes) - len(stoich))
            for g, s in zip(genes, stoich):
                locus = _normalise_locus(g)
                kb.setdefault(locus, []).append(
                    ComplexMembership(
                        complex_name=str(name),
                        stoichiometry=s,
                        init_count=init_count,
                        pdb_structures=pdbs,
                    )
                )
        return cls(gene_to_memberships=kb)

    def is_subunit(self, locus_tag: str) -> bool:
        return locus_tag in self.gene_to_memberships

    def active_memberships(self, locus_tag: str) -> list[ComplexMembership]:
        return [m for m in self.gene_to_memberships.get(locus_tag, [])
                if m.is_active]

    def all_locus_tags(self) -> set[str]:
        return set(self.gene_to_memberships.keys())


# Module-level singleton so per-predict overhead is zero.
_KB_SINGLETON: Optional[ComplexAssemblyKB] = None


def _get_kb(path: Optional[Path] = None) -> ComplexAssemblyKB:
    global _KB_SINGLETON
    if _KB_SINGLETON is None or path is not None:
        _KB_SINGLETON = ComplexAssemblyKB.load(path)
    return _KB_SINGLETON


@dataclass
class ComplexAssemblyDetector:
    """Predict essentiality from complex-subunit membership alone.

    Trajectory-agnostic: ``detect_for_gene`` accepts a ``Trajectory``
    only to match the ensemble interface, and ignores it.

    When the gene is a subunit of at least one ``is_active`` complex
    (init_count > 0), predicts TRANSLATION_STALL with confidence 0.9
    (the failure mode doesn't strictly fit — it's really a complex-
    assembly failure — but TRANSLATION_STALL is the closest existing
    FailureMode for ribosomal / translation-factor essentials that
    dominate this detector's TPs).

    Genes not in any known complex get NONE with zero confidence;
    the caller should combine with a trajectory-based detector.
    """

    kb: ComplexAssemblyKB = field(default_factory=_get_kb)
    confidence: float = 0.9

    def detect_for_gene(
        self,
        locus_tag: str,
        ko: Optional[Trajectory] = None,   # ignored
    ) -> tuple[FailureMode, Optional[float], float, str]:
        memberships = self.kb.active_memberships(locus_tag)
        if not memberships:
            return FailureMode.NONE, None, 0.0, "cx_none"
        # Build the evidence string from the smallest-init-count
        # complex the gene is a subunit of — that's the one most
        # likely to hit zero first under knockout.
        memberships.sort(key=lambda m: m.init_count)
        ev = "cx[" + ",".join(
            f"{m.complex_name}x{m.stoichiometry}(init={m.init_count})"
            for m in memberships[:3]
        ) + ("]" if len(memberships) <= 3 else ",...]")
        return (FailureMode.TRANSLATION_STALL, 0.0, self.confidence, ev)


def evaluate_against_breuer(
    kb: Optional[ComplexAssemblyKB] = None,
    breuer_csv: Optional[Path] = None,
) -> dict:
    """Offline sanity check: how well does the 'in-any-active-complex'
    rule match Breuer 2019 labels?

    Returns {tp, fp, tn, fn, unknown, mcc, precision, recall,
    covered_essentials, total_essentials}.
    """
    import csv
    if kb is None:
        kb = _get_kb()
    if breuer_csv is None:
        breuer_csv = (
            Path(__file__).resolve().parents[2]
            / "memory_bank" / "data" / "syn3a_essentiality_breuer2019.csv"
        )
    labels: dict[str, str] = {}
    with open(breuer_csv) as f:
        r = csv.DictReader(f)
        for row in r:
            labels[row["locus_tag"]] = row["essentiality"]

    # Positive prediction = in any active complex.
    tp = fp = tn = fn = unknown = 0
    for locus, label in labels.items():
        positive = bool(kb.active_memberships(locus))
        true_essential = label == "Essential"
        if label == "Essential":
            if positive: tp += 1
            else: fn += 1
        elif label == "Nonessential":
            if positive: fp += 1
            else: tn += 1
        else:   # Quasiessential / unknown
            unknown += 1

    # MCC on the two-class {Essential, Nonessential} subset.
    from math import sqrt
    num = tp * tn - fp * fn
    den_sq = (tp + fp) * (tp + fn) * (tn + fp) * (tn + fn)
    mcc = (num / sqrt(den_sq)) if den_sq > 0 else 0.0
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    total_essentials = sum(1 for v in labels.values() if v == "Essential")
    return {
        "tp": tp, "fp": fp, "tn": tn, "fn": fn, "unknown": unknown,
        "mcc": mcc,
        "precision": precision,
        "recall": recall,
        "covered_essentials": tp,
        "total_essentials": total_essentials,
    }


__all__ = [
    "ComplexAssemblyDetector",
    "ComplexAssemblyKB",
    "ComplexMembership",
    "evaluate_against_breuer",
]
