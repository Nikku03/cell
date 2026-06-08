"""Essential-gene pattern analysis using data already in this repo.

Three minimal Mycoplasma genomes with essentiality labels make a clean
test case: tiny genomes (~500-1500 genes each), high essential rates
(50-85%), well-studied in the literature.

Data sources (all in repo, no Drive/NCBI required):
  syn3a   memory_bank/data/syn3a_essentiality_breuer2019.csv
          memory_bank/data/syn3a_gene_table.csv
          memory_bank/data/syn3a_gene_class_features.csv (pre-computed
            length, GRAVY, pI, aromaticity, TM, AA composition, ...)
          memory_bank/data/multiorg_essentiality/sequences.fasta
  mgen    memory_bank/data/multiorg_essentiality/_xs_cache/L43967.2.gff3
          memory_bank/data/multiorg_essentiality/_xs_cache/protein_sequence.fasta
          memory_bank/data/multiorg_essentiality/_xs_cache/mg_essentiality_basis.xlsx
  mpne    memory_bank/data/multiorg_essentiality/_xs_cache/NC_000912.gff
          memory_bank/data/multiorg_essentiality/_xs_cache/mycoplasma.g-1.fasta
          memory_bank/data/multiorg_essentiality/_xs_cache/goldsets.csv

Computes per gene:
  STRUCTURAL    length_nt, length_aa, strand, relative position
  NEIGHBOR      intergenic distance prev/next, same-strand prev/next,
                operon-like flag
  PROTEIN       GRAVY (hydrophobicity), pI, aromaticity, net charge,
                amino-acid composition fractions (20 cols)
  ANNOTATION    product description, function class (syn3a only)
  TEXT          tokenised product words for enrichment

Then per-feature analysis (essential vs non-essential):
  numeric    -> mean +/- std for each class, Cohen's d, Mann-Whitney p,
                 univariate AUC
  text       -> word-enrichment (essential vs non-essential rates,
                 ratio, n_observed). Restricted to words appearing
                 >=5 times.
  function   -> proportion of essentials per category (syn3a)

Pooled across all three organisms + per-organism. Output:
  outputs/essential_repo_per_gene.csv          master table
  outputs/essential_repo_pooled_numeric.csv    per-feature stats
  outputs/essential_repo_word_enrichment.csv   product-word stats
  outputs/essential_repo_summary.md            readable report
"""
from __future__ import annotations
import argparse, math, re, sys
from collections import Counter, defaultdict
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent

# ---- amino-acid hydrophobicity (Kyte-Doolittle) and pKa values for pI ----
KD = {"A":1.8,"R":-4.5,"N":-3.5,"D":-3.5,"C":2.5,"E":-3.5,"Q":-3.5,"G":-0.4,
      "H":-3.2,"I":4.5,"L":3.8,"K":-3.9,"M":1.9,"F":2.8,"P":-1.6,"S":-0.8,
      "T":-0.7,"W":-0.9,"Y":-1.3,"V":4.2}
PKA_NTERM, PKA_CTERM = 9.69, 2.34
PKA_SIDE = {"K":10.5,"R":12.4,"H":6.0,"D":3.65,"E":4.25,"C":8.33,"Y":10.07}
AROMATIC = set("FWY")
AAS = "ACDEFGHIKLMNPQRSTVWY"

# stopwords for description enrichment
STOPWORDS = set("a an and are as at be by for from has have in is it its of on or "
                "that the their these this to with putative hypothetical protein "
                "uncharacterized predicted family domain containing related dna rna "
                "small large subunit type system component conserved member "
                "n-terminal c-terminal binding ".split())


def gravy(seq: str) -> float:
    seq = [a for a in seq if a in KD]
    return sum(KD[a] for a in seq) / len(seq) if seq else float("nan")


def aromaticity(seq: str) -> float:
    seq = [a for a in seq if a in KD]
    return sum(1 for a in seq if a in AROMATIC) / len(seq) if seq else float("nan")


def isoelectric(seq: str) -> float:
    """Compute pI via charge-zero search (binary)."""
    n = Counter(a for a in seq if a in KD)
    lo, hi = 0.0, 14.0
    def charge_at(pH):
        pos = 1/(1+10**(pH-PKA_NTERM))
        for aa, pka in (("K",PKA_SIDE["K"]),("R",PKA_SIDE["R"]),("H",PKA_SIDE["H"])):
            pos += n.get(aa,0) / (1+10**(pH-pka))
        neg = 1/(1+10**(PKA_CTERM-pH))
        for aa, pka in (("D",PKA_SIDE["D"]),("E",PKA_SIDE["E"]),
                          ("C",PKA_SIDE["C"]),("Y",PKA_SIDE["Y"])):
            neg += n.get(aa,0) / (1+10**(pka-pH))
        return pos - neg
    for _ in range(50):
        mid = (lo+hi)/2; c = charge_at(mid)
        if abs(c) < 1e-3: return mid
        if c > 0: lo = mid
        else: hi = mid
    return (lo+hi)/2


def aa_composition(seq: str) -> dict:
    seq = [a for a in seq if a in KD]
    n = len(seq); c = Counter(seq)
    return {f"aa_{a}": c.get(a,0)/n if n else 0 for a in AAS}


def parse_fasta(path: Path) -> dict[str, str]:
    out = {}; cur = None; buf = []
    with open(path) as f:
        for line in f:
            if line.startswith(">"):
                if cur is not None: out[cur] = "".join(buf)
                cur = line[1:].strip(); buf = []
            else:
                buf.append(line.strip())
    if cur is not None: out[cur] = "".join(buf)
    return out


def parse_gff_cds(path: Path) -> list[dict]:
    """Robust to two GFF styles:
       - modern: CDS rows carry locus_tag/gene/product directly
       - older RefSeq: locus_tag is on the 'gene' parent row;
                       CDS rows have Parent=<gene_id>
    Builds the join when needed."""
    gene_attrs = {}   # gene_id -> {locus_tag, gene, ...}
    cds_rows = []
    with open(path) as f:
        for line in f:
            if line.startswith("#") or not line.strip(): continue
            p = line.rstrip("\n").split("\t")
            if len(p) < 9: continue
            d = {}
            for kv in p[8].split(";"):
                if "=" in kv:
                    k, v = kv.split("=",1); d[k] = v
            if p[2] == "gene":
                gid = d.get("ID","")
                if gid:
                    gene_attrs[gid] = {"locus_tag": d.get("locus_tag",""),
                                        "gene":      d.get("gene","")}
            elif p[2] == "CDS":
                cds_rows.append((p, d))
    feats = []
    for p, d in cds_rows:
        lt = d.get("locus_tag",""); gene = d.get("gene","")
        if not lt and d.get("Parent","") in gene_attrs:
            ga = gene_attrs[d["Parent"]]
            lt = ga["locus_tag"]; gene = gene or ga["gene"]
        feats.append({
            "seqid":      p[0],
            "start":      int(p[3]),
            "end":        int(p[4]),
            "strand":     p[6],
            "locus_tag":  lt,
            "gene":       gene,
            "product":    d.get("product",""),
            "protein_id": d.get("protein_id",""),
        })
    return feats


# ---- stats helpers ----
def mann_whitney_p(a, b):
    import numpy as np
    a = np.array([x for x in a if x == x])
    b = np.array([x for x in b if x == x])
    if len(a) < 5 or len(b) < 5: return float("nan")
    c = np.concatenate([a, b])
    r = c.argsort().argsort() + 1
    u1 = r[:len(a)].sum() - len(a)*(len(a)+1)/2
    u = min(u1, len(a)*len(b) - u1)
    mu = len(a)*len(b)/2
    s = math.sqrt(len(a)*len(b)*(len(a)+len(b)+1)/12)
    if s == 0: return 1.0
    return math.erfc(abs((u-mu)/s)/math.sqrt(2))


def cohens_d(a, b):
    import numpy as np
    a = np.array([x for x in a if x == x])
    b = np.array([x for x in b if x == x])
    if len(a) < 2 or len(b) < 2: return float("nan")
    s = math.sqrt(((len(a)-1)*a.var(ddof=1) + (len(b)-1)*b.var(ddof=1)) /
                   (len(a)+len(b)-2))
    return (a.mean()-b.mean())/s if s else float("nan")


def auc_uni(scores, labels):
    import numpy as np
    s = np.array(scores, dtype=float); y = np.array(labels)
    m = ~np.isnan(s); s, y = s[m], y[m]
    if y.sum() == 0 or y.sum() == len(y): return float("nan")
    o = np.argsort(-s); y2 = y[o]
    np_pos = int(y2.sum()); nn = len(y2) - np_pos
    rs = np.where(y2 == 1)[0].sum() + np_pos
    return 1 - (rs - np_pos*(np_pos+1)/2) / (np_pos*nn)


def fisher_log_odds(a_ess, n_ess, a_non, n_non):
    """log-odds of being-essential | word present, vs absent."""
    p1 = (a_ess + 0.5) / (n_ess + 1)
    p0 = (a_non + 0.5) / (n_non + 1)
    return math.log(p1/(1-p1)) - math.log(p0/(1-p0))


# ---- per-organism loaders ----
def load_syn3a() -> list[dict]:
    import pandas as pd
    ess = pd.read_csv(REPO_ROOT/"memory_bank/data/syn3a_essentiality_breuer2019.csv")
    table = pd.read_csv(REPO_ROOT/"memory_bank/data/syn3a_gene_table.csv")
    feats = pd.read_csv(REPO_ROOT/"memory_bank/data/syn3a_gene_class_features.csv")
    seqs = parse_fasta(REPO_ROOT/"memory_bank/data/multiorg_essentiality/sequences.fasta")
    # essentiality: 'Essential' -> 1; otherwise 0 (Non-essential, Quasi-essential)
    ess_map = dict(zip(ess.locus_tag, ess.essentiality.str.lower().eq("essential").astype(int)))
    func_map = dict(zip(ess.locus_tag, ess.primary_function))
    rows = []
    f_idx = feats.set_index("locus_tag")
    table_sorted = table.sort_values("start_1based").reset_index(drop=True)
    contig_len = int(table_sorted["end"].max())
    for i, r in table_sorted.iterrows():
        lt = r["locus_tag"]
        if lt not in ess_map: continue
        # neighbor intergenic
        if i > 0:
            ig_prev = int(r["start_1based"] - table_sorted["end"].iloc[i-1] - 1)
            ss_prev = int(table_sorted["strand"].iloc[i-1] == r["strand"])
        else:
            ig_prev, ss_prev = None, None
        if i < len(table_sorted)-1:
            ig_next = int(table_sorted["start_1based"].iloc[i+1] - r["end"] - 1)
            ss_next = int(table_sorted["strand"].iloc[i+1] == r["strand"])
        else:
            ig_next, ss_next = None, None
        # protein sequence
        gene_str = r["gene_name"] if pd.notna(r["gene_name"]) else ""
        seq_key = f"syn3a|{lt}|{gene_str}".rstrip("|")
        seq = seqs.get(seq_key, "")
        # use precomputed features where available
        rec = {
            "organism":      "syn3a",
            "locus_tag":     lt,
            "gene":          gene_str,
            "product":       r["product"] if pd.notna(r["product"]) else "",
            "length_nt":     int(r["length_bp"]),
            "length_aa":     int(r["length_bp"]/3 - 1),
            "strand":        1 if r["strand"] == "+" else -1,
            "rel_position":  r["start_1based"] / contig_len,
            "intergenic_prev": ig_prev, "intergenic_next": ig_next,
            "same_strand_prev": ss_prev, "same_strand_next": ss_next,
            "operon_prev":   int(ig_prev is not None and ig_prev < 50 and ss_prev == 1),
            "essential":     ess_map[lt],
            "function_class":func_map.get(lt, ""),
        }
        if seq:
            rec["gravy"] = gravy(seq)
            rec["pI"]    = isoelectric(seq)
            rec["aromaticity"] = aromaticity(seq)
            comp = aa_composition(seq)
            rec.update(comp)
        elif lt in f_idx.index:
            for col in ["gravy","pI","aromaticity"] + [f"aa_{a}" for a in AAS]:
                if col in f_idx.columns:
                    v = f_idx.loc[lt, col]
                    if isinstance(v, (int, float)) and v == v: rec[col] = float(v)
        rows.append(rec)
    return rows


def load_with_gff(organism: str, gff_path: Path, fasta_path: Path,
                  labels: dict[str, int]) -> list[dict]:
    """For mgen and mpne: parse GFF, FASTA, intersect with labels."""
    feats = parse_gff_cds(gff_path)
    feats.sort(key=lambda r: (r["seqid"], r["start"]))
    seqs_raw = parse_fasta(fasta_path)
    # FASTA header schemes vary -- index by both first-token and by locus_tag substr
    seqs = {}
    for h, s in seqs_raw.items():
        first = h.split()[0]
        seqs[first] = s
        # mge:MG_001 -> MG_001 ; NC_000912\tMPN001\t... -> MPN001
        for token in re.split(r"[\s\t|:]+", h):
            if re.match(r"^(MG_|MPN)\d+", token):
                seqs[token] = s
    contig_lens = {f["seqid"]: 0 for f in feats}
    for f in feats:
        if f["end"] > contig_lens[f["seqid"]]: contig_lens[f["seqid"]] = f["end"]
    rows = []
    for i, f in enumerate(feats):
        lt = f["locus_tag"]
        if not lt or lt not in labels: continue
        prev_f = feats[i-1] if i > 0 and feats[i-1]["seqid"] == f["seqid"] else None
        next_f = feats[i+1] if i+1 < len(feats) and feats[i+1]["seqid"] == f["seqid"] else None
        ig_prev = (f["start"] - prev_f["end"] - 1) if prev_f else None
        ig_next = (next_f["start"] - f["end"] - 1) if next_f else None
        ss_prev = int(prev_f["strand"] == f["strand"]) if prev_f else None
        ss_next = int(next_f["strand"] == f["strand"]) if next_f else None
        # protein sequence: try locus_tag, protein_id, gene
        seq = ""
        for key in (f["locus_tag"], f["protein_id"], f["gene"]):
            if key in seqs:
                seq = seqs[key]; break
        length_nt = f["end"] - f["start"] + 1
        rec = {
            "organism":      organism,
            "locus_tag":     lt,
            "gene":          f["gene"],
            "product":       f["product"],
            "length_nt":     length_nt,
            "length_aa":     int(length_nt/3 - 1),
            "strand":        1 if f["strand"] == "+" else -1,
            "rel_position":  f["start"] / contig_lens[f["seqid"]],
            "intergenic_prev": ig_prev, "intergenic_next": ig_next,
            "same_strand_prev": ss_prev, "same_strand_next": ss_next,
            "operon_prev":   int(ig_prev is not None and ig_prev < 50 and ss_prev == 1),
            "essential":     labels[lt],
        }
        if seq:
            rec["gravy"] = gravy(seq)
            rec["pI"]    = isoelectric(seq)
            rec["aromaticity"] = aromaticity(seq)
            rec.update(aa_composition(seq))
        rows.append(rec)
    return rows


def load_mgen() -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(REPO_ROOT/"memory_bank/data/multiorg_essentiality/_xs_cache/mg_essentiality_basis.xlsx")
    ws = wb["Sheet1"]
    labels = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        gene = row[0]; exp = row[4]  # 'Exp' column
        if gene and exp in ("Y", "N"):
            labels[str(gene).strip()] = 1 if exp == "Y" else 0
    return load_with_gff(
        "mgen",
        REPO_ROOT/"memory_bank/data/multiorg_essentiality/_xs_cache/L43967.2.gff3",
        REPO_ROOT/"memory_bank/data/multiorg_essentiality/_xs_cache/protein_sequence.fasta",
        labels)


def load_mpne() -> list[dict]:
    import pandas as pd
    g = pd.read_csv(REPO_ROOT/"memory_bank/data/multiorg_essentiality/_xs_cache/goldsets.csv",
                     sep="\t")
    labels = dict(zip(g.gene, g["class"].eq("E").astype(int)))
    return load_with_gff(
        "mpne",
        REPO_ROOT/"memory_bank/data/multiorg_essentiality/_xs_cache/NC_000912.gff",
        REPO_ROOT/"memory_bank/data/multiorg_essentiality/_xs_cache/mycoplasma.g-1.fasta",
        labels)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--out-dir", type=Path,
                   default=REPO_ROOT / "outputs" / "essential_analysis")
    args = p.parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    try:
        import pandas as pd
    except ImportError as e:
        print(f"ERROR: {e}", file=sys.stderr); return 1

    print("Loading syn3a ..."); rs = load_syn3a(); print(f"  {len(rs)} rows")
    print("Loading mgen ...");  rg = load_mgen();  print(f"  {len(rg)} rows")
    print("Loading mpne ...");  rp = load_mpne();  print(f"  {len(rp)} rows")
    df = pd.DataFrame(rs + rg + rp)
    print(f"\nMaster: {len(df):,} genes  ({df.organism.nunique()} organisms)")
    print(df.groupby("organism").agg(n=("essential","count"),
                                       n_ess=("essential","sum"),
                                       pct_ess=("essential","mean")).to_string())
    df.to_csv(args.out_dir / "essential_repo_per_gene.csv", index=False)

    # numeric features
    base_num = ["length_nt","length_aa","strand","rel_position",
                "intergenic_prev","intergenic_next",
                "same_strand_prev","same_strand_next","operon_prev",
                "gravy","pI","aromaticity"]
    aa_cols = [f"aa_{a}" for a in AAS]
    numeric_feats = [c for c in base_num + aa_cols if c in df.columns]

    print(f"\n=== POOLED numeric features: essential vs non-essential ===")
    print(f"  {'feature':<22s} {'ess_mean':>10s} {'non_mean':>10s} {'d':>6s} "
          f"{'p_MW':>9s} {'AUC':>5s}")
    stats = []
    em = df.essential == 1
    for f in numeric_feats:
        a = df.loc[em, f].dropna().tolist()
        b = df.loc[~em, f].dropna().tolist()
        if len(a) < 5 or len(b) < 5: continue
        d = cohens_d(a, b); pv = mann_whitney_p(a, b)
        auc = auc_uni(df[f].tolist(), df.essential.tolist())
        em_v = sum(a)/len(a); nm_v = sum(b)/len(b)
        stats.append({"feature":f,"ess_mean":em_v,"non_mean":nm_v,
                       "cohen_d":d,"p_mannwhitney":pv,"auc":auc,
                       "n_ess":len(a),"n_non":len(b)})
    # sort by |d|
    stats.sort(key=lambda r:-abs(r["cohen_d"]) if r["cohen_d"]==r["cohen_d"] else 0)
    for r in stats[:25]:
        print(f"  {r['feature']:<22s} {r['ess_mean']:>10.4f} {r['non_mean']:>10.4f} "
              f"{r['cohen_d']:>+5.2f} {r['p_mannwhitney']:>9.2e} {r['auc']:>5.3f}")
    pd.DataFrame(stats).to_csv(args.out_dir / "essential_repo_pooled_numeric.csv", index=False)

    # per-organism strongest separators
    print(f"\n=== PER-ORGANISM: top 5 most separating features ===")
    per_org_rows = []
    for org in df.organism.unique():
        sub = df[df.organism == org]
        em_o = sub.essential == 1
        if em_o.sum() < 20 or (~em_o).sum() < 20: continue
        rows = []
        for f in numeric_feats:
            a = sub.loc[em_o, f].dropna().tolist()
            b = sub.loc[~em_o, f].dropna().tolist()
            if len(a) < 3 or len(b) < 3: continue
            d = cohens_d(a, b)
            auc = auc_uni(sub[f].tolist(), sub.essential.tolist())
            rows.append((f, d, auc))
        rows.sort(key=lambda x:-abs(x[1]) if x[1]==x[1] else 0)
        print(f"\n  {org}  (n_ess={int(em_o.sum())}/{len(sub)})")
        for f, d, auc in rows[:8]:
            print(f"    {f:<22s}  d={d:>+5.2f}   AUC={auc:.3f}")
            per_org_rows.append({"organism":org,"feature":f,"cohen_d":d,"auc":auc})
    pd.DataFrame(per_org_rows).to_csv(args.out_dir / "essential_repo_per_org.csv", index=False)

    # function-class breakdown for syn3a
    print(f"\n=== syn3a: essentiality rate by function class ===")
    syn = df[df.organism == "syn3a"]
    fc = syn.groupby("function_class").agg(
        n=("essential","count"),
        n_ess=("essential","sum"),
        pct_ess=("essential","mean")).sort_values("pct_ess", ascending=False)
    print(fc.to_string())
    fc.to_csv(args.out_dir / "essential_repo_syn3a_function.csv")

    # word enrichment in product descriptions
    print(f"\n=== POOLED: words enriched in essential gene descriptions ===")
    def tokens(s):
        if not s: return []
        s = re.sub(r"[^A-Za-z0-9 -]"," ", str(s)).lower()
        return [t for t in s.split() if len(t) >= 3 and t not in STOPWORDS]
    ess_w, non_w = Counter(), Counter()
    ess_n_genes = int((df.essential==1).sum())
    non_n_genes = int((df.essential==0).sum())
    seen_w_ess = Counter()      # genes-containing-word, ess
    seen_w_non = Counter()
    for _, r in df.iterrows():
        ws = set(tokens(r["product"]))
        for t in ws:
            (seen_w_ess if r.essential==1 else seen_w_non)[t] += 1
            (ess_w if r.essential==1 else non_w)[t] += 1
    rows = []
    for w in set(ess_w) | set(non_w):
        n_e = seen_w_ess.get(w, 0); n_n = seen_w_non.get(w, 0)
        if n_e + n_n < 5: continue
        pe = n_e / ess_n_genes; pn = n_n / non_n_genes
        rows.append({"word": w, "n_ess_genes": n_e, "n_non_genes": n_n,
                      "pct_ess": pe*100, "pct_non": pn*100,
                      "enrichment": (pe+1e-6)/(pn+1e-6),
                      "log_odds": fisher_log_odds(n_e, ess_n_genes, n_n, non_n_genes)})
    wdf = pd.DataFrame(rows)
    wdf_e = wdf.sort_values("log_odds", ascending=False)
    print(f"\n  TOP enriched in essential:")
    print(f"  {'word':<25s} {'n_ess':>6s} {'n_non':>6s} {'ess%':>6s} {'non%':>6s} {'enrich':>7s}")
    for _, r in wdf_e.head(25).iterrows():
        print(f"  {r.word:<25s} {int(r.n_ess_genes):>6d} {int(r.n_non_genes):>6d} "
              f"{r.pct_ess:>5.1f}% {r.pct_non:>5.1f}% {r.enrichment:>6.2f}x")
    print(f"\n  TOP DEPLETED in essential (enriched in NON-essential):")
    for _, r in wdf_e.tail(15).iterrows():
        print(f"  {r.word:<25s} {int(r.n_ess_genes):>6d} {int(r.n_non_genes):>6d} "
              f"{r.pct_ess:>5.1f}% {r.pct_non:>5.1f}% {r.enrichment:>6.2f}x")
    wdf_e.to_csv(args.out_dir / "essential_repo_word_enrichment.csv", index=False)

    # ---- markdown report ----
    md = []
    md.append(f"# Essential-gene patterns across 3 minimal Mycoplasma genomes\n\n")
    md.append(f"**Organisms analyzed:** syn3a (JCVI-syn3.0), mgen "
              f"(M. genitalium G37), mpne (M. pneumoniae M129)  \n")
    md.append(f"**Total labeled genes:** {len(df):,}  \n")
    md.append(f"**Pooled essential rate:** {df.essential.mean()*100:.1f}%\n\n")
    md.append("## Per-organism counts\n\n")
    md.append("| organism | n_genes | n_essential | % essential |\n|---|---|---|---|\n")
    for org in df.organism.unique():
        sub = df[df.organism == org]
        md.append(f"| {org} | {len(sub)} | {int(sub.essential.sum())} | "
                  f"{sub.essential.mean()*100:.1f}% |\n")

    md.append("\n## Top features distinguishing essential from non-essential (pooled)\n\n")
    md.append("Sorted by |Cohen's d| (effect size). AUC = univariate classifier.\n\n")
    md.append("| feature | ess_mean | non_mean | Cohen's d | AUC | p (MW) |\n")
    md.append("|---|---|---|---|---|---|\n")
    for r in stats[:20]:
        md.append(f"| {r['feature']} | {r['ess_mean']:.3f} | {r['non_mean']:.3f} | "
                  f"{r['cohen_d']:+.2f} | {r['auc']:.3f} | {r['p_mannwhitney']:.1e} |\n")

    md.append("\n## syn3a: essentiality by function class\n\n")
    md.append("| function class | n | n_essential | % essential |\n|---|---|---|---|\n")
    for fc_name, fc_row in fc.iterrows():
        md.append(f"| {fc_name} | {int(fc_row.n)} | {int(fc_row.n_ess)} | "
                  f"{fc_row.pct_ess*100:.1f}% |\n")

    md.append("\n## Words enriched in essential gene product descriptions\n\n")
    md.append("| word | n_ess_genes | n_non_genes | ess% | non% | enrichment |\n")
    md.append("|---|---|---|---|---|---|\n")
    for _, r in wdf_e.head(20).iterrows():
        md.append(f"| {r.word} | {int(r.n_ess_genes)} | {int(r.n_non_genes)} | "
                  f"{r.pct_ess:.1f}% | {r.pct_non:.1f}% | {r.enrichment:.2f}x |\n")

    md.append("\n## Words depleted in essential (enriched in non-essential)\n\n")
    md.append("| word | n_ess_genes | n_non_genes | ess% | non% | enrichment |\n")
    md.append("|---|---|---|---|---|---|\n")
    for _, r in wdf_e.tail(15).iterrows():
        md.append(f"| {r.word} | {int(r.n_ess_genes)} | {int(r.n_non_genes)} | "
                  f"{r.pct_ess:.1f}% | {r.pct_non:.1f}% | {r.enrichment:.2f}x |\n")

    (args.out_dir / "essential_repo_summary.md").write_text("".join(md))
    print(f"\nwrote markdown summary -> {args.out_dir/'essential_repo_summary.md'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
